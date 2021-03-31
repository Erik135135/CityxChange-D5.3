##### Function for writing csv files
def read_file(excel,periods_on = True, periods=288):
    """
    Convert data from an excel file into a csv file. In case of having timeseries we need to make sure the first
    column contains the date as "05.12.2019  13:50:00". The hour will be rounded to XX:00, XX:05, XX:10, XX:15...

    The csv file will contain the number of periods introduced starting from the current time.
    """

    import csv
    import openpyxl
    import os
    import datetime as dt

    ##### Functions for datetime formatting
    def time_mod(time, delta, epoch=None):
        if epoch is None:
            epoch = dt.datetime(1970, 1, 1, tzinfo=time.tzinfo)
        return (time - epoch) % delta

    def time_round(time, delta, epoch=None):
        mod = time_mod(time, delta, epoch)
        if mod < (delta / 2):
            return time - mod
        return time + (delta - mod)

    tm = dt.datetime.now() # get actual datetime
    delta = dt.timedelta(minutes=1) # round it to 5 min
    date = str(time_round(tm,delta)) # get the string to match it with the excel sheet data
    # print("The real datetime is: ", tm)
    # print("The new datetime rounded to 5 minutes is: ", time_round(tm, delta))

    # get actual directory to go to the "Data" folder
    p = os.getcwd()
    os.chdir(r"Data")
    path_data = os.getcwd() # path for the Data folder
    path_excel = path_data + "\\" + excel # give us the path for the Excel Sheet

    # get access to the excel sheet
    data_excel = openpyxl.load_workbook((path_excel)) # import workbook
    sheetnames = data_excel.sheetnames.copy() # take all the sheetnames from the workbook


    for s in sheetnames:
        index_count = 1
        # print("Creating csv for " + s)
        sh = data_excel[str(s)] # take the sheetnames
        with open(f'{s}.csv', 'w', newline="") as f: # generates a new csv file with the name of the sheet
            c = csv.writer((f))
            if periods_on==True:
                sh.insert_cols(1)
                for row in sh.iter_rows(min_row=1, max_row=1):  # write the heading
                    row[0].value = "Index"
                    c.writerow([cell.value for cell in row])

                for i in range(2, sh.max_row): # look for the values of periods in the first column

                    # print("sh.cell type: ", type(sh.cell(row=i, column=1).value))
                    # print(sh.cell(row=i, column=1).value)
                    # print(time_round(sh.cell(row=i, column=1).value, delta))
                    # print("date value: ", date)
                    # print(value_cell)
                    # print("date type: ", type(date))
                    # print(type(str(value_cell)))
                    # print(sh.cell(row=i, column=2).value)
                    value_cell = time_round(sh.cell(row=i, column=2).value, delta)

                    if str(value_cell) == date:
                        # print("The model is run for: ", value_cell, "=", date)
                        for row in sh.iter_rows(min_row=i, max_row=periods+i):
                            row[0].value = index_count
                            index_count = index_count + 1
                            row[1].value = time_round(row[1].value,delta) # round the data to the nearest 5 min
                            # print(row[1].value)
                            c.writerow([cell.value for cell in row])
                    else:
                        continue

            else:
                for row in sh.iter_rows():  # it gets the rows from the last row
                    # print("\n")
                    # print(cell.value for cell in row)
                    c.writerow([cell.value for cell in row])

            # sh.delete_cols(1, amount=1)

    os.chdir(p)

# read_file("trial.xlsx", periods_on=True)
# read_file('Data_periods_v1.xlsx', periods_on=True, periods=24) # contains parameters with periods as index
# # read_file('Data_heatpumps.xlsx', periods_on=False) # contains parameters with heat pump number as index
# read_file("CollectEl.xlsx", periods_on=False)