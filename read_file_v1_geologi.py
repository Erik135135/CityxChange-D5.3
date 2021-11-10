##### Function for writing csv files
def read_file_v1(excel,periods_on = True, periods=288):
    """
    Convert data from an excel file into a csv file. In case of having timeseries we need to make sure the first
    column contains the date as "05.12.2019  13:50:00". The hour will be rounded to XX:00, XX:05, XX:10, XX:15...

    The csv file will contain the number of periods introduced starting from the current time.

    Adapted for forecast update
    """

    import csv
    import openpyxl
    import os
    import datetime
    import pandas as pd

    def roundTime(dt=None, dateDelta=datetime.timedelta(minutes=5)): # Reading data and sorting them into 5 minutes intervals
        """Round a datetime object to a multiple of a timedelta
        dt : datetime.datetime object, default now.
        dateDelta : timedelta object, we round to a multiple of this, default 1 minute.
        Author: Thierry Husson 2012 - Use it as you want but don't blame me.
                Stijn Nevens 2014 - Changed to use only datetime objects as variables
        """
        roundTo = dateDelta.total_seconds()

        if dt == None:
            dt = datetime.datetime.now()
        seconds = (dt - dt.min).seconds
        # // is a floor division, not a comment on following line:
        rounding = (seconds + roundTo / 2) // roundTo * roundTo
        return dt + datetime.timedelta(0, rounding - seconds, -dt.microsecond)

    # tm = datetime.datetime.now() # get actual datetime
    # delta = datetime.timedelta(minutes=5) # round it to 5 min
    # date = str(roundTime(tm)) # get the string to match it with the excel sheet data

    # get actual directory to go to the "Data" folder
    p = os.getcwd()
    os.chdir(r"Data")
    path_data = os.getcwd() # path for the Data folder
    path_excel = path_data + "\\" + excel # give us the path for the Excel Sheet

    # get access to the excel sheet
    data_excel = openpyxl.load_workbook((path_excel)) # import workbook
    sheetnames = data_excel.sheetnames.copy() # take all the sheetnames from the workbook

    for s in sheetnames:
        index_count = 1
        sh = data_excel[str(s)] # take the sheetnames
        with open(f'{s}.csv', 'w', newline="") as f: # generates a new csv file with the name of the sheet
            c = csv.writer((f))
            if periods_on==True:
                sh.insert_cols(1)
                for row in sh.iter_rows(min_row=1, max_row=1):  # write the headings
                    row[0].value = "Index"
                    row[1].value = "Time"
                    c.writerow([cell.value for cell in row])

                for row in sh.iter_rows(min_row=2, max_row=sh.max_row):
                    row[0].value = index_count
                    index_count = index_count + 1
                    # print(sh.cell(row=index_count, column=2).value)
                    row[1].value = roundTime(sh.cell(row=index_count, column=2).value) # round the data to the nearest 5 min
                    c.writerow([cell.value for cell in row])

            else:
                for row in sh.iter_rows():  # it gets the rows from the last row
                    # print("\n")
                    # print(cell.value for cell in row)
                    c.writerow([cell.value for cell in row])

    os.chdir(p)
